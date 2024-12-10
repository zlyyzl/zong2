import streamlit as st
import pandas as pd
import base64
import sqlite3
import hashlib
from PIL import Image
import matplotlib.pyplot as plt
import pickle
from sklearn.calibration import calibration_curve
from sklearn.metrics import brier_score_loss, accuracy_score, recall_score, precision_score, f1_score, roc_auc_score, roc_curve
from sklearn.metrics import roc_curve, auc, accuracy_score, recall_score, precision_score, f1_score 
from pycaret import classification
from pycaret.classification import setup, get_logs,create_model, predict_model, tune_model, save_model,load_model
import shap
shap.initjs()
import joblib
from sklearn.preprocessing import OneHotEncoder
import numpy as np
import streamlit.components.v1 as components
import openpyxl
import os

def get_db_connection():
    conn = sqlite3.connect('users.db')
    conn.row_factory = sqlite3.Row
    return conn

def initialize_database():
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def register_user(username, password):
    conn = get_db_connection()
    try:
        conn.execute("INSERT INTO users (username, password) VALUES (?, ?)",
                     (username, hash_password(password)))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def validate_user(username, password):
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE username = ? AND password = ?",
                        (username, hash_password(password))).fetchone()
    conn.close()
    return user is not None

def login_page():
    st.title("User Login")
    
    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login")

        if submitted:
            if validate_user(username, password):
                st.session_state['is_logged_in'] = True
                st.success("Login successful!")
                st.experimental_rerun()
            else:
                st.error("Invalid username or password.")

def register_page():
    st.title("User Registration")

    with st.form("register_form"):
        new_username = st.text_input("New Username")
        new_password = st.text_input("New Password", type="password")
        register_button = st.form_submit_button("Register")

        if register_button:
            if register_user(new_username, new_password):
                st.success("Registration successful! Please log in.")
            else:
                st.error("That username is already taken. Please try another username.")
                
def prediction_page():
    import streamlit as st

    # Set the page title
    st.set_page_config(page_title="My Application", layout="centered")


    page = st.sidebar.selectbox("Select a Page", ["Home Page", "Prediction"])

    if page == "Home Page":
        st.title("Welcome to Functional outcome prediction App for patients with posterior circulation large vessel occlusion after mechanical thrombectomy")
    
        st.header("Summary")
        st.write("""
            This application aims to predict functional outcome in  patients with posterior circulation large vessel occlusion following mechanical thrombectomy，thus facilitates informed clinical judgment, supports personalized treatment and follow-up plans, and establishes realistic treatment expectations.
        """)

        st.header("Main Features")
        features = [
            "✔️ Implementation of preoperative, intraoperative, and postoperative prediction models to dynamically update predictions of functional outcomes.",
            "✔️ Support for batch predictions of functional outcomes for multiple patients.",
            "✔️ Ability to predict outcomes for patients with missing variable values.",
            "✔️ Facilitation of the interpretation of how the model provides personalized predictions for specific cases.",
            "✔️ Consideration of changing environments with automatic deployment of updated prediction models."
        ]
        for feature in features:
            st.write(feature)

        st.header("How to Use")
        st.markdown("""
            To the left, is a dropdown main menu for navigating to each page in the present App:<br><br>
            
            &bull; **Home Page:** We are here!<br>
            
            &bull; **Prediction:** Overview of the prediction section.<br>
            
            &bull; **Preoperative_number:** Manage preoperative predictions by inputting the necessary data.<br>
            
            &bull; **Preoperative_batch:** Process preoperative batch predictions by uploading a file.<br>
            
            &bull; **Perioperative_number:** Manage perioperative predictions by inputting the necessary data.<br>
            
            &bull; **Perioperative_batch:** Process perioperative batch predictions by uploading a file.<br>
            
            &bull; **Postoperative_number:** Manage postoperative predictions by inputting the necessary data.<br>
            
            &bull; **Postoperative_batch:** Process postoperative batch predictions by uploading a file.<br>
            
            """, unsafe_allow_html=True)

        pdf_file_path = r"123.pdf"
        if os.path.exists(pdf_file_path):
            st.markdown("Click here to download the manual for more detailed usage instructions:")
            with open(pdf_file_path, "rb") as f:
                st.download_button(label="User Manual.pdf",data=f, file_name="User Manual.pdf",mime="application/pdf" )
        else:
            st.error("指定的文件不存在，请检查文件路径。")
        
        st.header("Contact Us")
        st.write("""
            If you have any questions, please contact the support team:
            - Email: 2894683001@qq.com
            
        """)
        
        st.header("Useful Links")
        st.markdown(
        """
        An app designed to predict functional outcomes for patients with anterior circulation large vessel occlusion following mechanical thrombectomy:
         - [Visit our partner site](https://zhelvyao-123-60-anterior.streamlit.app/)
         """)

        st.markdown(
            f"""
            <style>
                .appview-container .main .block-container{{
                    max-width: {1150}px;
                    padding-top: {5}rem;
                    padding-right: {10}rem;
                    padding-left: {10}rem;
                    padding-bottom: {5}rem;
                }}
                .reportview-container .main {{
                    color: white;
                    background-color: black;
                }}
            </style>
            """,
            unsafe_allow_html=True
        )
    
        image = Image.open('it.tif')
        st.image(image, use_column_width=True)

    elif page == "Prediction":
        st.title('Functional outcome prediction App for patients with posterior circulation large vessel occlusion after mechanical thrombectomy')
    
        model = joblib.load('tuned_rf_pre_BUN.pkl')
        model2 = load_model('tuned_rf_pre_BUN_model')
        model3 = joblib.load('tuned_rf_intra_BUN.pkl')
        model4 = load_model('tuned_rf_intra_BUN_model')
        model5 = joblib.load('tuned_rf_post_BUN.pkl')
        model6 = load_model('tuned_rf_post_BUN_model')
        def st_shap(plot, height=None):
            shap_html = f"<head>{shap.getjs()}</head><body>{plot.html()}</body>"
            components.html(shap_html, height=height)

        prediction_type = st.sidebar.selectbox(
            "How would you like to predict?",
            ("Preoperative_number", "Preoperative_batch", "Intraoperative_number", "Intraoperative_batch", 
             "Postoperative_number", "Postoperative_batch")
         )

        if prediction_type == "Preoperative_number":
            st.subheader("Preoperative Number Prediction")
            st.write("This section will manage preoperative predictions by inputting the necessary data.Please fill in the blanks with corresponding data. After that,click on the Predict button at the bottom to see the prediction of the classifier. ")
    
            NIHSS = st.number_input('NIHSS', min_value = 4,max_value = 38,value = 10) 
            GCS= st.number_input('GCS', min_value = 0,max_value = 15 ,value = 10) 
            pre_eGFR = st.number_input('pre_eGFR', min_value = 10.00,max_value = 250.00,value = 111.5)
            pre_glucose = st.number_input('pre_glucose', min_value = 2.50, max_value = 25.00, value = 7.78)
            PC_ASPECTS = st.number_input('PC_ASPECTS', min_value = 0.0,max_value = 10.0,value = 8.0)
            Age = st.number_input('Age', min_value = 0,max_value = 120,value = 60)
            pre_BUN = st.number_input('pre_BUN', min_value = 0.00,max_value = 30.00,value = 10.20)
            output=""

            features = { 
                'NIHSS': NIHSS, 
                'GCS': GCS, 
                'pre_eGFR': pre_eGFR,
                'pre_glucose': pre_glucose, 
                'PC_ASPECTS': PC_ASPECTS, 
                'Age': Age, 
                'pre_BUN': pre_BUN
                  }

            print(features) 

            input_df = pd.DataFrame([features]) 
            print(input_df) 
    
            if st.button('Predict'): 
                output = model.predict_proba(input_df)[:,1] 
                explainer = shap.Explainer(model)
                shap_values = explainer.shap_values(input_df)
                st.write(' Based on feature values, predicted possibility of good functional outcome is '+ str(output))
                st_shap(shap.force_plot(explainer.expected_value[1], shap_values[1],input_df))
                shap_df = pd.DataFrame({
                    'Feature': input_df.columns,
                    'SHAP Value': shap_values[1].flatten() 
                 })

                st.write("SHAP values for each feature:")
                st.dataframe(shap_df)
            label = st.selectbox('Outcome for Learning', [0, 1])
            if st.button('Add Data for Learning'): 
                model.fit(input_df, [label])
                st.success("New data has been added to the model for continuous learning!")

        elif prediction_type == "Preoperative_batch":
            st.subheader("Preoperative Batch Prediction")
            def plot_roc_curve(y_true, y_scores): 
                fpr, tpr, thresholds = roc_curve(y_true, y_scores) 
                roc_auc = auc(fpr, tpr)

                plt.figure(figsize=(10, 6)) 
                plt.plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc) 
                plt.plot([0, 1], [0, 1], color='red', linestyle='--') 
                plt.xlim([0.0, 1.0]) 
                plt.ylim([0.0, 1.05]) 
                plt.xlabel('False Positive Rate') 
                plt.ylabel('True Positive Rate')
                plt.title('Receiver Operating Characteristic (ROC)') 
                plt.legend(loc='lower right') 
                plt.grid() 
                st.pyplot(plt) 
            st.write("This section will handle preoperative batch predictions.Please click on the link to download the form and fill in the corresponding data. After that, click on the Browse files button to upload file for prediciton, you can see the prediction of the classifier at the bottom. This page supports batch prediction of the outcome of multiple patients at one time, and can predict the outcome of patients with missing values.")

            csv_exporter=openpyxl.Workbook()
            sheet=csv_exporter.active
            sheet.cell(row=1,column=1).value='NIHSS'
            sheet.cell(row=1,column=2).value='GCS'
            sheet.cell(row=1,column=3).value='pre_eGFR'
            sheet.cell(row=1,column=4).value='pre_glucose'
            sheet.cell(row=1,column=5).value='PC_ASPECTS'
            sheet.cell(row=1,column=6).value='Age'
            sheet.cell(row=1,column=7).value='pre_BUN'
            csv_exporter.save('for predictions.csv')

            data = open('for_predictions.csv', 'rb').read() 
            b64 = base64.b64encode(data).decode('UTF-8') 
            href = f'<a href="data:file/data;base64,{b64}" download="for_predictions.csv">Download csv file</a>'
            st.markdown(href, unsafe_allow_html=True) 
            csv_exporter.close() 

            file_upload = st.file_uploader("Upload CSV file for predictions", type=["csv"]) 

            if file_upload is not None: 
                try: 
                    data = pd.read_csv(file_upload, sep=',', error_bad_lines=False) 

                    if 'MRSI' in data.columns: 
                        y_true = data['MRSI'].values 
                        predictions = model2.predict_proba(data)[:, 1] 
                        predictions_df = pd.DataFrame(predictions, columns=['Predictions']) 
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions_df 
                        result_file_path = 'predictions_with_results.csv' 
                        result_data.to_csv(result_file_path, index=False) 
 
                        with open(result_file_path, 'rb') as f: 
                            output_data = f.read()
                            b64 = base64.b64encode(output_data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                            st.markdown(download_link, unsafe_allow_html=True)

                        add_data = st.selectbox('Outcome for Learning', [0, 1])
                        if st.button('Add Data for Learning'): 
                            X = data.drop(columns=['MRSI']) 
                            y = data['MRSI'] 
                            model2.fit(X, y) 
                            st.success("New data has been added to the model for continuous learning!")

                        def plot_combined_graphs(y_true, y_scores):
                            fig, axs = plt.subplots(1, 2, figsize=(14, 6))
                            fpr, tpr, _ = roc_curve(y_true, y_scores)
                            roc_auc = auc(fpr, tpr)
                            axs[0].plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc)
                            axs[0].plot([0, 1], [0, 1], color='red', linestyle='--')
                            axs[0].set_xlim([0.0, 1.0])
                            axs[0].set_ylim([0.0, 1.05])
                            axs[0].set_xlabel('False Positive Rate')
                            axs[0].set_ylabel('True Positive Rate')
                            axs[0].set_title('Receiver Operating Characteristic (ROC)')
                            axs[0].legend(loc='lower right')
                            axs[0].grid()

                            prob_true, prob_pred = calibration_curve(y_true, y_scores, n_bins=10)
                            axs[1].plot(prob_pred, prob_true, marker='o', label='Calibrated Model', color='b')
                            axs[1].plot([0, 1], [0, 1], linestyle='--', label='Perfectly Calibrated', color='r')
                            axs[1].set_title('Brier Score Calibration Plot')
                            axs[1].set_xlabel('Mean Predicted Probability')
                            axs[1].set_ylabel('Fraction of Positives')
                            axs[1].set_xlim([0, 1])
                            axs[1].set_ylim([0, 1])
                            axs[1].legend()
                            axs[1].grid()
                            st.pyplot(fig)

                        if len(data) >= 10:
                            y_pred = (predictions > 0.5).astype(int)
                            accuracy = accuracy_score(y_true, y_pred)
                            recall = recall_score(y_true, y_pred)
                            precision = precision_score(y_true, y_pred)
                            f1 = f1_score(y_true, y_pred)
                            roc_auc = auc(*roc_curve(y_true, predictions)[:2])

                            st.write(f"Accuracy: {accuracy:.2f}")
                            st.write(f"Recall: {recall:.2f}")
                            st.write(f"Precision: {precision:.2f}")
                            st.write(f"F1 Score: {f1:.2f}")
                            st.write(f"AUC: {roc_auc:.2f}")
                            brier_score = brier_score_loss(y_true, predictions)
                            st.write(f"Brier Score: {brier_score:.2f}")

                            plot_combined_graphs(y_true, predictions)

                        else:
                            st.warning("Not enough samples for ROC curve plotting. Please upload at least 10 samples.") 

                            st.write(predictions_df) 
                    else:                      
                        predictions = model2.predict_proba(data)[:,1] 
                        predictions = pd.DataFrame(predictions,columns = ['Predictions'])
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions 
                        result_file_path = 'predictions_with_results.csv'
                        result_data.to_csv(result_file_path, index=False)
                        with open(result_file_path, 'rb') as f:
                            data = f.read()
                            b64 = base64.b64encode(data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>' 
                            st.markdown(download_link, unsafe_allow_html=True)

                except Exception as e: 
                    st.error(f"Error reading the CSV file: {e}") 
                    
        elif prediction_type == "Intraoperative_number":
            st.subheader("Intraoperative Number Prediction")
            st.write("This section will handle intraoperative number predictions.please fill in the blanks with corresponding data. After that,click on the Predict button at the bottom to see the prediction of the classifier.")

            NIHSS = st.number_input('NIHSS', min_value = 4,max_value = 38,value = 10) 
            GCS= st.number_input('GCS', min_value = 0,max_value = 15 ,value = 10) 
            pre_eGFR = st.number_input('pre_eGFR', min_value = 10.00,max_value = 250.00,value = 111.5)
            PC_ASPECTS = st.number_input('PC_ASPECTS', min_value = 0.0,max_value = 10.0,value = 8.0)
            Age = st.number_input('Age', min_value = 0,max_value = 120,value = 60)
            pre_BUN = st.number_input('pre_BUN', min_value = 0.20,max_value = 30.00,value = 3.20)
            procedural_time = st.number_input('procedural time', min_value=0.00, max_value=350.00, value=60.00)          
            output = ""

    # 这里不需要将 procedural_time 放入列表中
            features = { 
                'NIHSS': NIHSS, 
                'GCS': GCS, 
                'pre_eGFR': pre_eGFR, 
                'PC_ASPECTS': PC_ASPECTS,
                'Age': Age,
                'pre_BUN': pre_BUN,
                'procedural time': procedural_time                  
               
                 }

            print(features) 

            input_df = pd.DataFrame([features])

            print(input_df) 

    
            if st.button('Predict'): 
                output = model3.predict_proba(input_df)[:,1] 
                explainer = shap.Explainer(model3)
                shap_values = explainer.shap_values(input_df)
                st.write(' Based on feature values, predicted possibility of good functional outcome is '+ str(output))
                st_shap(shap.force_plot(explainer.expected_value[1], shap_values[1],input_df))
                shap_df = pd.DataFrame({
                    'Feature': input_df.columns,
                    'SHAP Value': shap_values[1].flatten() 
                 })
                st.write("SHAP values for each feature:")
                st.dataframe(shap_df)
        
            label = st.selectbox('Outcome for Learning', [0, 1])
            if st.button('Add Data for Learning'): 
                model3.fit(input_df, [label])
                st.success("New data has been added to the model for continuous learning!")

        elif prediction_type == "Intraoperative_batch":
            st.subheader("Intraoperative Batch Prediction")
            st.write("This section will handle intraoperative batch predictions.Please click on the link to download the form and fill in the corresponding data. After that, click on the Browse files button to upload file for prediciton, you can see the prediction of the classifier at the bottom. This page supports batch prediction of the outcome of multiple patients at one time, and can predict the outcome of patients with missing values.")
            def plot_roc_curve(y_true, y_scores): 
                fpr, tpr, thresholds = roc_curve(y_true, y_scores) 
                roc_auc = auc(fpr, tpr)

                plt.figure(figsize=(10, 6)) 
                plt.plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc) 
                plt.plot([0, 1], [0, 1], color='red', linestyle='--') 
                plt.xlim([0.0, 1.0]) 
                plt.ylim([0.0, 1.05]) 
                plt.xlabel('False Positive Rate') 
                plt.ylabel('True Positive Rate')
                plt.title('Receiver Operating Characteristic (ROC)') 
                plt.legend(loc='lower right') 
                plt.grid() 
                st.pyplot(plt) 

            csv_exporter=openpyxl.Workbook()
            sheet=csv_exporter.active
            sheet.cell(row=1,column=1).value='NIHSS'
            sheet.cell(row=1,column=2).value='GCS'
            sheet.cell(row=1,column=3).value='pre_eGFR'
            sheet.cell(row=1,column=4).value='PC_ASPECTS'
            sheet.cell(row=1,column=5).value='Age'
            sheet.cell(row=1,column=6).value='pre_BUN'
            sheet.cell(row=1,column=7).value='procedural time'          
            csv_exporter.save('for predictions.csv')

            data = open('for_predictions.csv', 'rb').read() 
            b64 = base64.b64encode(data).decode('UTF-8') 
            href = f'<a href="data:file/data;base64,{b64}" download="for_predictions.csv">Download csv file</a>'
            st.markdown(href, unsafe_allow_html=True) 
            csv_exporter.close() 

            file_upload = st.file_uploader("Upload CSV file for predictions", type=["csv"]) 

            if file_upload is not None: 
                try: 
                    data = pd.read_csv(file_upload, sep=',', error_bad_lines=False) 

                    if 'MRSI' in data.columns: 
                        y_true = data['MRSI'].values 
                        predictions = model4.predict_proba(data)[:, 1] 
                        predictions_df = pd.DataFrame(predictions, columns=['Predictions']) 
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions_df 
                        result_file_path = 'predictions_with_results.csv' 
                        result_data.to_csv(result_file_path, index=False) 
 
                        with open(result_file_path, 'rb') as f: 
                            output_data = f.read()
                            b64 = base64.b64encode(output_data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                            st.markdown(download_link, unsafe_allow_html=True)

                        add_data = st.selectbox('Outcome for Learning', [0, 1])
                
                        if st.button('Add Data for Learning'): 
                            X = data.drop(columns=['MRSI']) 
                            y = data['MRSI'] 
                            model4.fit(X, y) 
                            st.success("New data has been added to the model for continuous learning!")

                        def plot_combined_graphs(y_true, y_scores):
                            fig, axs = plt.subplots(1, 2, figsize=(14, 6))
                            fpr, tpr, _ = roc_curve(y_true, y_scores)
                            roc_auc = auc(fpr, tpr)
                            axs[0].plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc)
                            axs[0].plot([0, 1], [0, 1], color='red', linestyle='--')
                            axs[0].set_xlim([0.0, 1.0])
                            axs[0].set_ylim([0.0, 1.05])
                            axs[0].set_xlabel('False Positive Rate')
                            axs[0].set_ylabel('True Positive Rate')
                            axs[0].set_title('Receiver Operating Characteristic (ROC)')
                            axs[0].legend(loc='lower right')
                            axs[0].grid()

                            prob_true, prob_pred = calibration_curve(y_true, y_scores, n_bins=10)
                            axs[1].plot(prob_pred, prob_true, marker='o', label='Calibrated Model', color='b')
                            axs[1].plot([0, 1], [0, 1], linestyle='--', label='Perfectly Calibrated', color='r')
                            axs[1].set_title('Brier Score Calibration Plot')
                            axs[1].set_xlabel('Mean Predicted Probability')
                            axs[1].set_xlim([0, 1])
                            axs[1].set_ylim([0, 1])
                            axs[1].legend()
                            axs[1].grid()
                            st.pyplot(fig)

                        if len(data) >= 10:
                            y_pred = (predictions > 0.5).astype(int)
                            accuracy = accuracy_score(y_true, y_pred)
                            recall = recall_score(y_true, y_pred)
                            precision = precision_score(y_true, y_pred)
                            f1 = f1_score(y_true, y_pred)
                            roc_auc = auc(*roc_curve(y_true, predictions)[:2])

                            st.write(f"Accuracy: {accuracy:.2f}")
                            st.write(f"Recall: {recall:.2f}")
                            st.write(f"Precision: {precision:.2f}")
                            st.write(f"F1 Score: {f1:.2f}")
                            st.write(f"AUC: {roc_auc:.2f}")
                            brier_score = brier_score_loss(y_true, predictions)
                            st.write(f"Brier Score: {brier_score:.2f}")

                            plot_combined_graphs(y_true, predictions)
                        else:
                            st.warning("Not enough samples for ROC curve plotting. Please upload at least 10 samples.") 

                            st.write(predictions_df) 
                    else:                      
                        predictions = model4.predict_proba(data)[:,1] 
                        predictions = pd.DataFrame(predictions,columns = ['Predictions'])
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions 
                        result_file_path = 'predictions_with_results.csv'
                        result_data.to_csv(result_file_path, index=False)
                        with open(result_file_path, 'rb') as f:
                            data = f.read()
                            b64 = base64.b64encode(data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                            st.markdown(download_link, unsafe_allow_html=True)

                except Exception as e: 
                    st.error(f"Error reading the CSV file: {e}") 
            

        elif prediction_type == "Postoperative_number":
            st.subheader("Postoperative Number Prediction")
            st.write("This section will handle postoperative number predictions.please fill in the blanks with corresponding data. After that,click on the Predict button at the bottom to see the prediction of the classifier.")
            Age = st.number_input('Age', min_value = 0, max_value = 120, value = 60)
            GCS= st.number_input('GCS', min_value = 0,max_value = 15 ,value = 10)  
            PC_ASPECTS = st.number_input('PC_ASPECTS', min_value = 0.0,max_value = 10.0,value = 8.0)
            procedural_time = st.number_input('procedural time', min_value=0.00, max_value=350.00, value=60.00)
            post_eGFR = st.number_input('post_eGFR', min_value = 10.00,max_value = 250.00,value = 111.5) 
            post_NIHSS = st.number_input('post_NIHSS', min_value = 4,max_value = 38,value = 10) 
                            
            output=""
            features = { 
                'Age': Age, 
                'GCS': GCS, 
                'PC_ASPECTS': PC_ASPECTS,
                'procedural time': procedural_time, 
                'post_eGFR': post_eGFR,
                'post_NIHSS': post_NIHSS            
                
                  }

            print(features) 

            input_df = pd.DataFrame([features])

            print(input_df) 
    
            if st.button('Predict'): 
                output = model5.predict_proba(input_df)[:,1] 
                explainer = shap.Explainer(model5)
                shap_values = explainer.shap_values(input_df)
                st.write(' Based on feature values, predicted possibility of good functional outcome is '+ str(output))
                st_shap(shap.force_plot(explainer.expected_value[1], shap_values[1],input_df))
                shap_df = pd.DataFrame({
                    'Feature': input_df.columns,
                    'SHAP Value': shap_values[1].flatten() 
                 })
                st.write("SHAP values for each feature:")
                st.dataframe(shap_df)
        
            label = st.selectbox('Outcome for Learning', [0, 1])
            if st.button('Add Data for Learning'): 
                model5.fit(input_df, [label])
                st.success("New data has been added to the model for continuous learning!")

        elif prediction_type == "Postoperative_batch":
            st.subheader("Postoperative Batch Prediction")
            st.write("This section will handle postoperative batch predictions.Please click on the link to download the form and fill in the corresponding data. After that, click on the Browse files button to upload file for prediciton, you can see the prediction of the classifier at the bottom. This page supports batch prediction of the outcome of multiple patients at one time, and can predict the outcome of patients with missing values.")
            def plot_roc_curve(y_true, y_scores): 
                fpr, tpr, thresholds = roc_curve(y_true, y_scores) 
                roc_auc = auc(fpr, tpr)

                plt.figure(figsize=(10, 6)) 
                plt.plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc) 
                plt.plot([0, 1], [0, 1], color='red', linestyle='--') 
                plt.xlim([0.0, 1.0]) 
                plt.ylim([0.0, 1.05]) 
                plt.xlabel('False Positive Rate') 
                plt.ylabel('True Positive Rate')
                plt.title('Receiver Operating Characteristic (ROC)') 
                plt.legend(loc='lower right') 
                plt.grid() 
                st.pyplot(plt) 

            csv_exporter = openpyxl.Workbook() 
            sheet = csv_exporter.active 
            sheet.cell(row=1, column=1).value = 'Age'
            sheet.cell(row=1, column=2).value = 'GCS' 
            sheet.cell(row=1, column=3).value = 'PC_ASPECTS' 
            sheet.cell(row=1, column=4).value = 'procedural time' 
            sheet.cell(row=1, column=5).value = 'post_eGFR' 
            sheet.cell(row=1, column=6).value = 'post_NIHSS'      
            
            csv_exporter.save('for_predictions.csv') 

            data = open('for_predictions.csv', 'rb').read() 
            b64 = base64.b64encode(data).decode('UTF-8') 
            href = f'<a href="data:file/data;base64,{b64}" download="for_predictions.csv">Download csv file</a>'
            st.markdown(href, unsafe_allow_html=True) 
            csv_exporter.close() 

            file_upload = st.file_uploader("Upload CSV file for predictions", type=["csv"]) 

            if file_upload is not None: 
                try: 
                    data = pd.read_csv(file_upload, sep=',', error_bad_lines=False) 

                    if 'MRSI' in data.columns: 
                        y_true = data['MRSI'].values 
                        predictions = model6.predict_proba(data)[:, 1] 
                        predictions_df = pd.DataFrame(predictions, columns=['Predictions']) 
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions_df 
                        result_file_path = 'predictions_with_results.csv' 
                        result_data.to_csv(result_file_path, index=False) 
 
                        with open(result_file_path, 'rb') as f: 
                            output_data = f.read()
                            b64 = base64.b64encode(output_data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                            st.markdown(download_link, unsafe_allow_html=True)

                        add_data = st.selectbox('Outcome for Learning', [0, 1])
                
                        if st.button('Add Data for Learning'): 
                            X = data.drop(columns=['MRSI']) 
                            y = data['MRSI'] 
                            model6.fit(X, y) 
                            st.success("New data has been added to the model for continuous learning!")

                        def plot_combined_graphs(y_true, y_scores):
                            fig, axs = plt.subplots(1, 2, figsize=(14, 6))
                            fpr, tpr, _ = roc_curve(y_true, y_scores)
                            roc_auc = auc(fpr, tpr)
                            axs[0].plot(fpr, tpr, color='blue', label='ROC curve (area = %0.2f)' % roc_auc)
                            axs[0].plot([0, 1], [0, 1], color='red', linestyle='--')
                            axs[0].set_xlim([0.0, 1.0])
                            axs[0].set_ylim([0.0, 1.05])
                            axs[0].set_xlabel('False Positive Rate')
                            axs[0].set_ylabel('True Positive Rate')
                            axs[0].set_title('Receiver Operating Characteristic (ROC)')
                            axs[0].legend(loc='lower right')
                            axs[0].grid()

                            prob_true, prob_pred = calibration_curve(y_true, y_scores, n_bins=10)
                            axs[1].plot(prob_pred, prob_true, marker='o', label='Calibrated Model', color='b')
                            axs[1].plot([0, 1], [0, 1], linestyle='--', label='Perfectly Calibrated', color='r')
                            axs[1].set_title('Brier Score Calibration Plot')
                            axs[1].set_xlabel('Mean Predicted Probability')
                            axs[1].set_ylabel('Fraction of Positives')
                            axs[1].set_xlim([0, 1])
                            axs[1].set_ylim([0, 1])
                            axs[1].legend()
                            axs[1].grid()
                            st.pyplot(fig)

                        if len(data) >= 10:
                            y_pred = (predictions > 0.5).astype(int)
                            accuracy = accuracy_score(y_true, y_pred)
                            recall = recall_score(y_true, y_pred)
                            precision = precision_score(y_true, y_pred)
                            f1 = f1_score(y_true, y_pred)
                            roc_auc = auc(*roc_curve(y_true, predictions)[:2])

                            st.write(f"Accuracy: {accuracy:.2f}")
                            st.write(f"Recall: {recall:.2f}")
                            st.write(f"Precision: {precision:.2f}")
                            st.write(f"F1 Score: {f1:.2f}")
                            st.write(f"AUC: {roc_auc:.2f}")
                            brier_score = brier_score_loss(y_true, predictions)
                            st.write(f"Brier Score: {brier_score:.2f}")
                            plot_combined_graphs(y_true, predictions)

                            if roc_auc < 0.75:
                                st.write("AUC is below 0.75. Retraining with Class Incremental Random Forests.")
                    
                                X = data.drop(columns=['MRSI']) 
                                y = data['MRSI'] 
                                cifr_model.partial_fit(X, y)  
                        else:
                            st.warning("Not enough samples for ROC curve plotting. Please upload at least 10 samples.") 

                            st.write(predictions_df) 
                    else:                      
                        predictions = model6.predict_proba(data)[:,1] 
                        predictions = pd.DataFrame(predictions,columns = ['Predictions'])
                        st.write(predictions)
                        result_data = data.copy() 
                        result_data['Predictions'] = predictions 
                        result_file_path = 'predictions_with_results.csv'
                        result_data.to_csv(result_file_path, index=False)
                        with open(result_file_path, 'rb') as f:
                            data = f.read()
                            b64 = base64.b64encode(data).decode('UTF-8')
                            download_link = f'<a href="data:file/csv;base64,{b64}" download="predictions_with_results.csv">Download predictions with results</a>'
                            st.markdown(download_link, unsafe_allow_html=True)

                except Exception as e: 
                    st.error(f"Error reading the CSV file: {e}")

    else:  # Other Features
        st.title("Other Features")
        st.write("Here you can describe other features of your application.")
  
def main():
    initialize_database()  

    if 'is_logged_in' not in st.session_state:
        st.session_state['is_logged_in'] = False

    if st.session_state['is_logged_in']:
        prediction_page()  
    else:
        login_or_register = st.sidebar.selectbox("Select an action:", ("Login", "Register"), key="login_register_selectbox")
        if login_or_register == "Login":
            login_page()
        else:
            register_page()

if __name__ == "__main__":
    main()


